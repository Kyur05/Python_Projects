import sys
import xml.etree.ElementTree as ET
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
import physics_engine

print("=== South African TRH17 Geometric Design Tool ===")

# --- THE EXCEL PAINTER ---
def format_worksheet(ws):
    """Applies professional UI styling to an Excel worksheet."""
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                         top=Side(style='thin'), bottom=Side(style='thin'))
    header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")
    
    pass_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    pass_text = Font(color="006100")
    fail_text = Font(color="9C0006")

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = thin_border
            cell.alignment = center_align
            
            if cell.row == 1:
                cell.fill = header_fill
                cell.font = Font(bold=True)
            
            if isinstance(cell.value, str):
                if "PASS" in cell.value:
                    cell.fill = pass_fill
                    cell.font = pass_text
                elif "CRITICAL" in cell.value or "ERROR" in cell.value:
                    cell.fill = fail_fill
                    cell.font = fail_text

    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length: max_length = len(str(cell.value))
            except: pass
        ws.column_dimensions[col_letter].width = max_length + 3


# --- THE STREAMING AUTOMATION ENGINE ---
def process_landxml_to_excel(xml_file, speed, topo):
    print(f"\nProcessing {xml_file} using Memory-Optimized Streaming...")
    output_filename = "N2_Master_Geometric_Report.xlsx"

    wb = openpyxl.Workbook()
    
    # Sheet 1: Vertical
    ws_vert = wb.active
    ws_vert.title = "Vertical Checks"
    ws_vert.append(["PVI CH", "PVI Elev", "G1 (%)", "G2 (%)", "Length", "Curve Type", "K-Value", "BVC CH", "EVC CH", "Turn CH", "Turn Elev", "Req SSD (m)"])
    
    # Sheet 2: Horizontal
    ws_horiz = wb.create_sheet(title="Horizontal Checks")
    ws_horiz.append(["Element", "Radius (m)", "Length", "TRH17 Min R", "Req Superelev (%)", "Status", "20m Survey Chords"])

    try:
        vert_points = []
        curve_count = 1

        # The Memory-Safe Extractor
        context = ET.iterparse(xml_file, events=("end",))
        
        for event, element in context:
            clean_tag = element.tag.split('}')[-1]
            
            if clean_tag == 'Curve':
                radius_str = element.get('radius')
                length_str = element.get('length')
                
                if radius_str and length_str:
                    radius = abs(float(radius_str))
                    length = float(length_str)
                    
                    horiz_data = physics_engine.calculate_horizontal_parameters(speed, radius)
                    chords = physics_engine.calculate_setting_out_chords(radius, length)
                    
                    ws_horiz.append([
                        f"Curve {curve_count}", f"{radius:.2f}", f"{length:.2f}", 
                        horiz_data['R_min_required'], horiz_data['e_required_percent'], 
                        horiz_data['status'], ", ".join(map(str, chords))
                    ])
                    curve_count += 1
                element.clear() # Instantly wipe from RAM

            elif clean_tag in ['PVI', 'ParaCurve']:
                if element.text and element.text.strip():
                    data = element.text.strip().split()
                    chainage = float(data[0])
                    elev = float(data[1])
                    length = float(element.get('length', 0)) if clean_tag == 'ParaCurve' else 0
                    vert_points.append({'chainage': chainage, 'elev': elev, 'length': length, 'tag': clean_tag})
                element.clear() # Instantly wipe from RAM

        # Process Vertical Math
        for i in range(1, len(vert_points) - 1):
            prev_pt = vert_points[i-1]
            curr_pt = vert_points[i]
            next_pt = vert_points[i+1]

            dx1 = curr_pt['chainage'] - prev_pt['chainage']
            dy1 = curr_pt['elev'] - prev_pt['elev']
            g1 = (dy1 / dx1 * 100) if dx1 != 0 else 0

            dx2 = next_pt['chainage'] - curr_pt['chainage']
            dy2 = next_pt['elev'] - curr_pt['elev']
            g2 = (dy2 / dx2 * 100) if dx2 != 0 else 0

            ssd = physics_engine.calculate_ssd(speed, g1)

            if curr_pt['tag'] == 'ParaCurve':
                v_data = physics_engine.calculate_vertical_geometry(curr_pt['chainage'], curr_pt['elev'], g1, g2, curr_pt['length'])
                ws_vert.append([
                    f"{curr_pt['chainage']:.2f}", f"{curr_pt['elev']:.2f}", f"{g1:.2f}", f"{g2:.2f}", 
                    curr_pt['length'], v_data['Curve_Type'], v_data['K_Value'], 
                    v_data['BVC_CH'], v_data['EVC_CH'], v_data['Turn_CH'], v_data['Turn_Elev'], ssd
                ])
            else:
                ws_vert.append([
                    f"{curr_pt['chainage']:.2f}", f"{curr_pt['elev']:.2f}", f"{g1:.2f}", f"{g2:.2f}", 
                    "-", "Tangent", "-", "-", "-", "-", "-", ssd
                ])

        # Apply UI Polish to both sheets
        format_worksheet(ws_vert)
        format_worksheet(ws_horiz)

        wb.save(output_filename)
        print(f"Advanced Master Report Successfully Generated: {output_filename}")

    except FileNotFoundError:
        print(f"ERROR: Could not find '{xml_file}'.")

if __name__ == "__main__":
    # If Dynamo is sending data, it will now send 3 things: The XML Path, Speed, and Topo
    if len(sys.argv) > 3:
        xml_path = sys.argv[1]
        dynamo_speed = int(sys.argv[2])
        dynamo_topo = sys.argv[3]
        process_landxml_to_excel(xml_path, dynamo_speed, dynamo_topo)
    else:
        # Fallback if you just click Play in VS Code manually
        process_landxml_to_excel("road_export.xml", 120, "Rolling")